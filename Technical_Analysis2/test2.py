import pandas as pd
import time
import os
import openpyxl
from pandas.io.excel._base import read_excel
from openpyxl.drawing.image import Image

class test2_cls:
    def exe_test2(self, jongname, buydate, inc_20, inc_60, inc_120, dec_20, dec_60, dec_120):
        # 데이터 가져오기
        data = pd.read_excel('F:/JusikData/API/Analy_Technical/Step2/juga_'+jongname+'_'+buydate+'.xlsx', engine='openpyxl')
        data_volpro = pd.read_excel('F:/JusikData/API/Analy_Technical/Step3/volpro_'+jongname+'_'+buydate+'.xlsx', engine='openpyxl')
        data_acquisition = pd.read_excel('F:/JusikData/API/Analy_Technical/Step3/acquisition_'+jongname+'_'+buydate+'.xlsx', engine='openpyxl')
        data_trends = pd.read_excel('F:/JusikData/API/Analy_Technical/Step3/trends_'+jongname+'_'+buydate+'.xlsx', engine='openpyxl')
        data_inclination = pd.read_excel('F:/JusikData/API/Analy_Technical/Step3/inclination_'+jongname+'_'+buydate+'.xlsx', engine='openpyxl')
        supply_buy_data = pd.read_excel('F:/JusikData/API/Analy_Technical/Step4/buy_'+jongname+'_'+buydate+'.xlsx', engine='openpyxl')
        supply_sell_data = pd.read_excel('F:/JusikData/API/Analy_Technical/Step4/sell_'+jongname+'_'+buydate+'.xlsx', engine='openpyxl')
        supply_sum_data = pd.read_excel('F:/JusikData/API/Analy_Technical/Step4/sum_'+jongname+'_'+buydate+'.xlsx', engine='openpyxl')
        
        # 당일 수급 구하기
        buy = supply_buy_data.loc[supply_buy_data['일자'] == int(buydate)]
        sell = supply_sell_data.loc[supply_sell_data['일자'] == int(buydate)]
        sum = supply_sum_data.loc[supply_sum_data['일자'] == int(buydate)]
        all1 = buy.append([sell, sum], sort=False, ignore_index=True)
        
        # 매물대 수급 구하기
        all2 = pd.DataFrame()
        for i in data_volpro.index:
            buy = supply_buy_data.loc[supply_buy_data['일자'] == data_volpro.iloc[i]['일자']]
            sell = supply_sell_data.loc[supply_sell_data['일자'] == data_volpro.iloc[i]['일자']]
            sum = supply_sum_data.loc[supply_sum_data['일자'] == data_volpro.iloc[i]['일자']]
            
            all2 = all2.append(buy, sort=False, ignore_index=True)
            all2 = all2.append(sell, sort=False, ignore_index=True)
            all2 = all2.append(sum, sort=False, ignore_index=True)
        
        # 매집봉 수급 구하기
        all3 = pd.DataFrame()
        if data_acquisition.empty :
            pass
        else:
            for i in data_acquisition.index:
                buy = supply_buy_data.loc[supply_buy_data['일자'] == data_acquisition.iloc[i]['일자']]
                buy2 = buy.copy()
                buy2['구분자'] = int(data_acquisition.iloc[i]['구분자'])
                buy2['구분'] = '매수'
                sell = supply_sell_data.loc[supply_sell_data['일자'] == data_acquisition.iloc[i]['일자']]
                sell2 = sell.copy()
                sell2['구분자'] = int(data_acquisition.iloc[i]['구분자'])
                sell2['구분'] = '매도'
                sum = supply_sum_data.loc[supply_sum_data['일자'] == data_acquisition.iloc[i]['일자']]
                sum2 = sum.copy()
                sum2['구분자'] = int(data_acquisition.iloc[i]['구분자'])
                sum2['구분'] = '순매수'
                
                all3 = all3.append(buy2, sort=False, ignore_index=True)
                all3 = all3.append(sell2, sort=False, ignore_index=True)
                all3 = all3.append(sum2, sort=False, ignore_index=True)
        
        # 누적 수급 구하기
        # 매물대 구간의 누적치를 구함(매물대는 제외, 현재일도 제외)
        arr_date_idx = []
        arr_date = []
        arr_class = []
        
        for i in data_volpro.index:
            dd = data_volpro.iloc[i]['일자']
            idx = supply_sum_data.loc[supply_sum_data['일자'] == dd].index[0]
            arr_date_idx.append(idx)
            arr_date.append(dd)
        arr_date_idx.sort() # 오름차순으로 정렬하기
        arr_date.sort()
        
        all4 = pd.DataFrame()
        for i in range(len(arr_date_idx)):
            if i == len(arr_date_idx) - 1:
                supply_data = supply_sum_data[arr_date_idx[i]+1:len(data)-1]
                
                supply_data_sum = pd.DataFrame(supply_data.sum(), columns=['0'])
                supply_data_sum = supply_data_sum.transpose()
                supply_data_sum['일자'] = arr_date[i]
                #print(supply_data_sum)
                
                all4 = all4.append(supply_data_sum, sort=False, ignore_index=True)          
            else : 
                supply_data = supply_sum_data[arr_date_idx[i]+1:len(data)-1] # 이부분만 수정하면 언제부터 언제까지 더할껀지 정할수있음
                
                supply_data_sum = pd.DataFrame(supply_data.sum(), columns=['0'])
                supply_data_sum = supply_data_sum.transpose()
                supply_data_sum['일자'] = arr_date[i]
                #print(supply_data_sum)
                
                all4 = all4.append(supply_data_sum, sort=False, ignore_index=True)     
        #print(all4)
        
        #--------------------------------------------------------------------------------------#
        ## 엑셀에 저장 ##
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook('F:/JusikData/API/Analy_Technical/종합/종목정리.xlsx')
        
        # 시트 지정 하기
        sheet = wb['전체정리']
        last_row = sheet.max_row

        # 표시할 데이터 뽑기
        시가 = data.iloc[-1]['시가']
        고가 = data.iloc[-1]['고가']
        저가 = data.iloc[-1]['저가']
        종가 = data.iloc[-1]['종가']
        추세가장기상 = data_trends.iloc[-1]['high_long']
        추세가장기하 = data_trends.iloc[-1]['low_long']
        추세가단기상 = data_trends.iloc[-1]['high_short']
        추세가단기하 = data_trends.iloc[-1]['low_short']
        장기울기상 = data_inclination.iloc[0]['장기울기상']
        장기울기하 = data_inclination.iloc[0]['장기울기하']
        단기울기상 = data_inclination.iloc[0]['단기울기상']
        단기울기하 = data_inclination.iloc[0]['단기울기하']
        
        if len(data_volpro) < 1:
            지금매물대 = "없음"
            지금강도 = "없음"
        else:
            지금매물대 = data_volpro.iloc[-1]['매물대']
            지금강도 = data_volpro.iloc[-1]['비율']
        
        # 만약 해소되지않은 매물대가 1개밖에 존재하지 않는다면
        if len(data_volpro) < 2:
            다음매물대 = "없음"
            다음강도 = "없음"
        else :
            다음매물대 = data_volpro.iloc[-2]['매물대']
            다음강도 = data_volpro.iloc[-2]['비율']
            
        # 매물대 자체 수급
        if len(data_volpro) < 2:
            매물대일자 = "없음"
        elif len(data_volpro) < 1:
            매물대일자 = data_volpro.iloc[-1]['일자']
        else : 
            if 고가>=지금매물대>종가 or 지금매물대 > 고가 :
                매물대일자 = data_volpro.iloc[-1]['일자']
            elif 종가>지금매물대 :
                매물대일자 = data_volpro.iloc[-2]['일자']
        
        if 매물대일자 == "없음":
            매수1등 = "없음"
            매수2등 = "없음"
            순매수1등 = "없음"
            순매수2등 = "없음"
            매집매수1등 = "없음"
            매집매수2등 = "없음"
            매집순매수1등 = "없음"
            매집순매수2등 = "없음"
            
            # 누적은 전체 누적으로 구하기
            supply_data = supply_sum_data
            supply_data_sum = pd.DataFrame(supply_data.sum(), columns=['0'])
            누적수급데이터 = supply_data_sum.transpose()
            del 누적수급데이터['일자']
            누적수급데이터 = 누적수급데이터.transpose()
            누적수급데이터.columns = ['매물대1누적'] # 컬럼명 지정하기
            누적수급데이터['누적랭크'] = 누적수급데이터['매물대1누적'].rank(method = 'first', ascending = False) # 랭크달기
            누적1등 = 누적수급데이터[누적수급데이터['누적랭크'] == 1.0].index[0]
            누적2등 = 누적수급데이터[누적수급데이터['누적랭크'] == 2.0].index[0]
            
        else:
            매물대수급데이터 = all2[all2['일자'] == 매물대일자] # 매물대1의 일자의 수급 가져오기
            del 매물대수급데이터['일자'] # 일자는 빼기
            매물대수급데이터 = 매물대수급데이터.transpose() # 행열 교환하기
            매물대수급데이터.columns = ['매수', '매도', '순매수'] # 컬럼명 지정하기
            매물대수급데이터['매수랭크'] = 매물대수급데이터['매수'].rank(method = 'first', ascending = False) # 랭크달기
            매물대수급데이터['매도랭크'] = 매물대수급데이터['매도'].rank(method = 'first', ascending = False) # 랭크달기
            매물대수급데이터['순매수랭크'] = 매물대수급데이터['순매수'].rank(method = 'first', ascending = False) # 랭크달기
            매수1등 = 매물대수급데이터[매물대수급데이터['매수랭크'] == 1.0].index[0]
            매수2등 = 매물대수급데이터[매물대수급데이터['매수랭크'] == 2.0].index[0]
            순매수1등 = 매물대수급데이터[매물대수급데이터['순매수랭크'] == 1.0].index[0]
            순매수2등 = 매물대수급데이터[매물대수급데이터['순매수랭크'] == 2.0].index[0]
            
            # 누적수급
            누적수급데이터 = all4[all4['일자'] == 매물대일자] # 매물대 일자의 수급 가져오기
            del 누적수급데이터['일자'] # 일자는 빼기
            누적수급데이터 = 누적수급데이터.transpose() # 행열 교환하기
            누적수급데이터.columns = ['매물대1누적'] # 컬럼명 지정하기
            누적수급데이터['누적랭크'] = 누적수급데이터['매물대1누적'].rank(method = 'first', ascending = False) # 랭크달기
            누적1등 = 누적수급데이터[누적수급데이터['누적랭크'] == 1.0].index[0]
            누적2등 = 누적수급데이터[누적수급데이터['누적랭크'] == 2.0].index[0]
            #print(누적수급데이터)
            
            # 매집봉수급
            # 해당 매물대 구간 매집봉수급
            if all3.empty :
                매집매수1등 = "없음"
                매집매수2등 = "없음"
                매집순매수1등 = "없음"
                매집순매수2등 = "없음"
            else:
                if 고가>=지금매물대>종가 or 지금매물대 > 고가 :
                    매물대구분자 = data_volpro.iloc[-1]['구분자']
                elif 종가>지금매물대 :
                    매물대구분자 = data_volpro.iloc[-2]['구분자']
                매집봉수급데이터 = all3[all3['구분자'] >= 매물대구분자]
                매집봉수급데이터매수 = 매집봉수급데이터[매집봉수급데이터['구분'] == '매수']
                del 매집봉수급데이터매수['일자'] # 일자는 빼기
                del 매집봉수급데이터매수['구분자']
                매집봉수급데이터매수.set_index('구분', inplace = True) # 구분은 인덱스로
                
                매집봉수급데이터매도 = 매집봉수급데이터[매집봉수급데이터['구분'] == '매도']
                del 매집봉수급데이터매도['일자'] # 일자는 빼기
                del 매집봉수급데이터매도['구분자']
                매집봉수급데이터매도.set_index('구분', inplace = True) # 구분은 인덱스로
                
                매집봉수급데이터순매수 = 매집봉수급데이터[매집봉수급데이터['구분'] == '순매수']
                del 매집봉수급데이터순매수['일자'] # 일자는 빼기
                del 매집봉수급데이터순매수['구분자']
                매집봉수급데이터순매수.set_index('구분', inplace = True) # 구분은 인덱스로
                
                # 매수, 매도, 순매수의 합을 데이터프레임으로 만들기        
                매집봉데이터all = pd.concat([매집봉수급데이터매수.sum(), 매집봉수급데이터매도.sum(), 매집봉수급데이터순매수.sum()], axis=1)
                매집봉데이터all.columns = ['매수', '매도', '순매수'] # 컬럼명 지정하기
                매집봉데이터all['매수랭크'] = 매집봉데이터all['매수'].rank(method = 'first', ascending = False) # 랭크달기
                매집봉데이터all['매도랭크'] = 매집봉데이터all['매도'].rank(method = 'first', ascending = False) # 랭크달기
                매집봉데이터all['순매수랭크'] = 매집봉데이터all['순매수'].rank(method = 'first', ascending = False) # 랭크달기
                
                매집매수1등 = 매집봉데이터all[매집봉데이터all['매수랭크'] == 1.0].index[0]
                매집매수2등 = 매집봉데이터all[매집봉데이터all['매수랭크'] == 2.0].index[0]
                매집순매수1등 = 매집봉데이터all[매집봉데이터all['순매수랭크'] == 1.0].index[0]
                매집순매수2등 = 매집봉데이터all[매집봉데이터all['순매수랭크'] == 2.0].index[0]
                
        # 당일수급
        오늘수급데이터 = all1.copy()
        del 오늘수급데이터['일자'] # 일자는 빼기
        오늘수급데이터 = 오늘수급데이터.transpose() # 행열 교환하기
        오늘수급데이터.columns = ['매수', '매도', '순매수'] # 컬럼명 지정하기
        오늘수급데이터['매수랭크'] = 오늘수급데이터['매수'].rank(method = 'first', ascending = False) # 랭크달기
        오늘수급데이터['매도랭크'] = 오늘수급데이터['매도'].rank(method = 'first', ascending = False) # 랭크달기
        오늘수급데이터['순매수랭크'] = 오늘수급데이터['순매수'].rank(method = 'first', ascending = False) # 랭크달기
        오늘매수1등 = 오늘수급데이터[오늘수급데이터['매수랭크'] == 1.0].index[0]
        오늘매수2등 = 오늘수급데이터[오늘수급데이터['매수랭크'] == 2.0].index[0]
        오늘순매수1등 = 오늘수급데이터[오늘수급데이터['순매수랭크'] == 1.0].index[0]
        오늘순매수2등 = 오늘수급데이터[오늘수급데이터['순매수랭크'] == 2.0].index[0]
        
        # 엑셀에 데이터 표시하기
        sheet.cell(last_row+1, 1, last_row-3)
        sheet.cell(last_row+1, 2, jongname)
        sheet.cell(last_row+1, 3, buydate)
        sheet.cell(last_row+1, 4, 시가)
        sheet.cell(last_row+1, 5, 고가)
        sheet.cell(last_row+1, 6, 저가)
        sheet.cell(last_row+1, 7, 종가)
        sheet.cell(last_row+1, 8, 추세가장기상)
        sheet.cell(last_row+1, 9, 추세가장기하)
        sheet.cell(last_row+1, 10, 추세가단기상)
        sheet.cell(last_row+1, 11, 추세가단기하)
        sheet.cell(last_row+1, 12, 장기울기상)
        sheet.cell(last_row+1, 13, 장기울기하)
        sheet.cell(last_row+1, 14, 단기울기상)
        sheet.cell(last_row+1, 15, 단기울기하)
        sheet.cell(last_row+1, 34, 지금매물대)
        sheet.cell(last_row+1, 36, 지금강도)
        sheet.cell(last_row+1, 37, 다음매물대)
        sheet.cell(last_row+1, 39, 다음강도)
        sheet.cell(last_row+1, 42, 매수1등)
        sheet.cell(last_row+1, 43, 매수2등)
        sheet.cell(last_row+1, 44, 매집매수1등)
        sheet.cell(last_row+1, 45, 매집매수2등)
        sheet.cell(last_row+1, 46, 오늘매수1등)
        sheet.cell(last_row+1, 47, 오늘매수2등)
        sheet.cell(last_row+1, 48, 순매수1등)
        sheet.cell(last_row+1, 49, 순매수2등)
        sheet.cell(last_row+1, 50, 누적1등)
        sheet.cell(last_row+1, 51, 누적2등)
        sheet.cell(last_row+1, 52, 매집순매수1등)
        sheet.cell(last_row+1, 53, 매집순매수2등)
        sheet.cell(last_row+1, 54, 오늘순매수1등)
        sheet.cell(last_row+1, 55, 오늘순매수2등)
        sheet.cell(last_row+1, 63, inc_20)
        sheet.cell(last_row+1, 64, inc_60)
        sheet.cell(last_row+1, 65, inc_120)
        sheet.cell(last_row+1, 66, dec_20)
        sheet.cell(last_row+1, 67, dec_60)
        sheet.cell(last_row+1, 68, dec_120)
        
        # 저장하기
        wb.save('F:/JusikData/API/Analy_Technical/종합/종목정리.xlsx')
        wb.close()
        
        
        
        
        
        
        
        
        
        '''
        # 데이터 프레임을 새로운 시트에 저장하기
        with pd.ExcelWriter('F:/JusikData/API/Analy_Technical/종합/'+jongname+'_'+buydate+'.xlsx', mode='a', engine='openpyxl') as writer:
            data.to_excel(writer, index=False, sheet_name = '수정주가')
            data_trends.to_excel(writer, index=False, sheet_name = '추세선')
            data_volpro.to_excel(writer, index=False, sheet_name = '매물대')
            data_acquisition.to_excel(writer, index=False, sheet_name = '매집봉')
            supply_buy_data.to_excel(writer, index=False, sheet_name = '매수')
            supply_sell_data.to_excel(writer, index=False, sheet_name = '매도')
            supply_sum_data.to_excel(writer, index=False, sheet_name = '순매수')
            all1.to_excel(writer, index=False, sheet_name = '오늘수급')
            all2.to_excel(writer, index=False, sheet_name = '매물대수급')
            all3.to_excel(writer, index=False, sheet_name = '매집봉수급')
            all4.to_excel(writer, index=False, sheet_name = '누적수급')
        '''
        
        '''         
        -> 매물대 ~ 현재일로 바꿀 코드 구간
        # 누적 수급 구하기
        # 매물대 구간의 누적치를 구함(매물대는 제외, 현재일도 제외)
        arr_date_idx = []
        arr_date = []
        for i in data_volpro.index:
            dd = data_volpro.iloc[i]['일자']
            idx = supply_sum_data.loc[supply_sum_data['일자'] == dd].index[0]
            arr_date_idx.append(idx)
            arr_date.append(dd)
        arr_date_idx.sort() # 오름차순으로 정렬하기
        arr_date.sort()
        
        all4 = pd.DataFrame()
        for i in range(len(arr_date_idx)):
            if i == len(arr_date_idx) - 1:
                supply_data = supply_sum_data[arr_date_idx[i]+1:len(data)-1]
                
                supply_data_sum = pd.DataFrame(supply_data.sum(), columns=['0'])
                supply_data_sum = supply_data_sum.transpose()
                supply_data_sum['일자'] = arr_date[i]
                #print(supply_data_sum)
                
                all4 = all4.append(supply_data_sum, sort=False, ignore_index=True)          
            else : 
                supply_data = supply_sum_data[arr_date_idx[i]+1:arr_date_idx[i+1]-1]
                
                supply_data_sum = pd.DataFrame(supply_data.sum(), columns=['0'])
                supply_data_sum = supply_data_sum.transpose()
                supply_data_sum['일자'] = arr_date[i]
                #print(supply_data_sum)
                
                all4 = all4.append(supply_data_sum, sort=False, ignore_index=True)
        '''
        '''
        #--------------------------------------------------------------------------------------#
        # 시고저종, 기울기, 추세선 가격
        data_add = []
        시가 = data.iloc[-1]['시가']
        고가 = data.iloc[-1]['고가']
        저가 = data.iloc[-1]['저가']
        종가 = data.iloc[-1]['종가']
        #print(시가,고가,저가,종가)
        
        장기울기상 = data_inclination.iloc[0]['장기울기상']
        장기울기하 = data_inclination.iloc[0]['장기울기하']
        단기울기상 = data_inclination.iloc[0]['단기울기상']
        단기울기하 = data_inclination.iloc[0]['단기울기하']
        #print(장기울기상,장기울기하,단기울기상,단기울기하)
        
        장추세가상 = data_trends.iloc[-1]['high_long']
        장추세가하 = data_trends.iloc[-1]['low_long']
        단추세가상 = data_trends.iloc[-1]['high_short']
        단추세가하 = data_trends.iloc[-1]['low_short']
        #print(장추세가상,장추세가하,단추세가상,단추세가하)
        
        data_add.append(jongname)
        data_add.append(buydate)
        data_add.append(장기울기상)
        data_add.append(장기울기하)
        data_add.append(단기울기상)
        data_add.append(단기울기하)
        
        # 증가, 감소율 구하기
        data_add.append(inc_20)
        data_add.append(inc_60)
        data_add.append(inc_120)
        data_add.append(dec_20)
        data_add.append(dec_60)
        data_add.append(dec_120)
        
        # 추가할 엑셀 열기
        data_in = openpyxl.load_workbook('F:/JusikData/API/Form/종합/분석양식.xlsx')
        sheet = data_in.active
        
        # 추가하기
        sheet.append(data_add)
        
        # 저장하기
        data_in.save('F:/JusikData/API/Analy_Technical/종합/'+jongname+'_'+buydate+'.xlsx')
        data_in.close()
        
        
        #--------------------------------------------------------------------------------------#
        ## # append # ##
        # 오늘 수급 append
        오늘수급데이터 = all1.copy()
        del 오늘수급데이터['일자'] # 일자는 빼기
        오늘수급데이터 = 오늘수급데이터.transpose() # 행열 교환하기
        오늘수급데이터.columns = ['매수', '매도', '순매수'] # 컬럼명 지정하기
        오늘수급데이터['매수랭크'] = 오늘수급데이터['매수'].rank(method = 'min', ascending = False) # 랭크달기
        오늘수급데이터['매도랭크'] = 오늘수급데이터['매도'].rank(method = 'min', ascending = False) # 랭크달기
        오늘수급데이터['순매수랭크'] = 오늘수급데이터['순매수'].rank(method = 'min', ascending = False) # 랭크달기
        오늘매수1등 = 오늘수급데이터[오늘수급데이터['매수랭크'] == 1.0].index[0]
        오늘매수량1 = 오늘수급데이터.loc[오늘매수1등]['매수']
        오늘매수2등 = 오늘수급데이터[오늘수급데이터['매수랭크'] == 2.0].index[0]
        오늘매수량2 = 오늘수급데이터.loc[오늘매수2등]['매수']
        오늘순매수1등 = 오늘수급데이터[오늘수급데이터['순매수랭크'] == 1.0].index[0]
        오늘순매수량1 = 오늘수급데이터.loc[오늘순매수1등]['순매수']
        오늘순매수2등 = 오늘수급데이터[오늘수급데이터['순매수랭크'] == 2.0].index[0]
        오늘순매수량2 = 오늘수급데이터.loc[오늘순매수2등]['순매수']
        오늘순매수3등 = 오늘수급데이터[오늘수급데이터['순매수랭크'] == 3.0].index[0]
        오늘순매수량3 = 오늘수급데이터.loc[오늘순매수3등]['순매수']
        
        data_add.append(오늘매수1등)
        data_add.append(오늘매수량1)
        data_add.append(오늘매수2등)
        data_add.append(오늘매수량2)
        data_add.append(오늘순매수1등)
        data_add.append(오늘순매수량1)
        data_add.append(오늘순매수2등)
        data_add.append(오늘순매수량2)
        data_add.append(오늘순매수3등)
        data_add.append(오늘순매수량3)
        
        #--------------------------------------------------------------------------------------#
        # for문으로 정리해보자
        # data_volpro2는 지금 매물대 크기순으로 정렬되어있음
        #print(data_volpro2)
        for i in data_volpro2.index:
            if len(data_volpro2) < 3 :
                if len(data_volpro2)-1 < i:
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    data_add.append("non")
                    
                    continue
                    
            if i == 3 :
                break
            
            # 매물대 정보
            매물대일자 = data_volpro2.iloc[i]['일자']
            매물대가격 = data_volpro2.iloc[i]['매물대']
            매물대거래량 = data_volpro2.iloc[i]['거래량']
            매물대유통주식 = data_volpro2.iloc[i]['유통주식수']
            매물대강도 = round((매물대거래량 / 매물대유통주식) * 100,2)
            매물대대비 = round(((매물대가격 - 종가) / 매물대가격) * 100,2)
            
            #-------------------------------#
            # 매물대 자체 수급
            매물대수급데이터 = all2[all2['일자'] == 매물대일자] # 매물대1의 일자의 수급 가져오기
            del 매물대수급데이터['일자'] # 일자는 빼기
            매물대수급데이터 = 매물대수급데이터.transpose() # 행열 교환하기
            매물대수급데이터.columns = ['매수', '매도', '순매수'] # 컬럼명 지정하기
            매물대수급데이터['매수랭크'] = 매물대수급데이터['매수'].rank(method = 'first', ascending = False) # 랭크달기
            매물대수급데이터['매도랭크'] = 매물대수급데이터['매도'].rank(method = 'first', ascending = False) # 랭크달기
            매물대수급데이터['순매수랭크'] = 매물대수급데이터['순매수'].rank(method = 'first', ascending = False) # 랭크달기
            #print(매물대수급데이터)
            매수1등 = 매물대수급데이터[매물대수급데이터['매수랭크'] == 1.0].index[0]
            매수량1 = 매물대수급데이터.loc[매수1등]['매수']
            매수2등 = 매물대수급데이터[매물대수급데이터['매수랭크'] == 2.0].index[0]
            매수량2 = 매물대수급데이터.loc[매수2등]['매수']
            순매수1등 = 매물대수급데이터[매물대수급데이터['순매수랭크'] == 1.0].index[0]
            순매수량1 = 매물대수급데이터.loc[순매수1등]['순매수']
            순매수2등 = 매물대수급데이터[매물대수급데이터['순매수랭크'] == 2.0].index[0]
            순매수량2 = 매물대수급데이터.loc[순매수2등]['순매수']
            순매수3등 = 매물대수급데이터[매물대수급데이터['순매수랭크'] == 3.0].index[0]
            순매수량3 = 매물대수급데이터.loc[순매수3등]['순매수']
            
            #-------------------------------#
            # 해당 매물대 구간 매집봉수급
            매집봉수급데이터 = all3[all3['구분자'] == data_volpro2.iloc[i]['구분자']]
            매집봉수급데이터매수 = 매집봉수급데이터[매집봉수급데이터['구분'] == '매수']
            del 매집봉수급데이터매수['일자'] # 일자는 빼기
            del 매집봉수급데이터매수['구분자']
            매집봉수급데이터매수.set_index('구분', inplace = True) # 구분은 인덱스로
            
            매집봉수급데이터매도 = 매집봉수급데이터[매집봉수급데이터['구분'] == '매도']
            del 매집봉수급데이터매도['일자'] # 일자는 빼기
            del 매집봉수급데이터매도['구분자']
            매집봉수급데이터매도.set_index('구분', inplace = True) # 구분은 인덱스로
            
            매집봉수급데이터순매수 = 매집봉수급데이터[매집봉수급데이터['구분'] == '순매수']
            del 매집봉수급데이터순매수['일자'] # 일자는 빼기
            del 매집봉수급데이터순매수['구분자']
            매집봉수급데이터순매수.set_index('구분', inplace = True) # 구분은 인덱스로
            
            # 매수, 매도, 순매수의 합을 데이터프레임으로 만들기        
            매집봉데이터all = pd.concat([매집봉수급데이터매수.sum(), 매집봉수급데이터매도.sum(), 매집봉수급데이터순매수.sum()], axis=1)
            매집봉데이터all.columns = ['매수', '매도', '순매수'] # 컬럼명 지정하기
            매집봉데이터all['매수랭크'] = 매집봉데이터all['매수'].rank(method = 'first', ascending = False) # 랭크달기
            매집봉데이터all['매도랭크'] = 매집봉데이터all['매도'].rank(method = 'first', ascending = False) # 랭크달기
            매집봉데이터all['순매수랭크'] = 매집봉데이터all['순매수'].rank(method = 'first', ascending = False) # 랭크달기
            
            매집매수1등 = 매집봉데이터all[매집봉데이터all['매수랭크'] == 1.0].index[0]
            매집매수량1 = 매집봉데이터all.loc[매집매수1등]['매수']
            매집매수2등 = 매집봉데이터all[매집봉데이터all['매수랭크'] == 2.0].index[0]
            매집매수량2 = 매집봉데이터all.loc[매집매수2등]['매수']
            매집순매수1등 = 매집봉데이터all[매집봉데이터all['순매수랭크'] == 1.0].index[0]
            매집순매수량1 = 매집봉데이터all.loc[매집순매수1등]['순매수']
            매집순매수2등 = 매집봉데이터all[매집봉데이터all['순매수랭크'] == 2.0].index[0]
            매집순매수량2 = 매집봉데이터all.loc[매집순매수2등]['순매수']
            매집순매수3등 = 매집봉데이터all[매집봉데이터all['순매수랭크'] == 3.0].index[0]
            매집순매수량3 = 매집봉데이터all.loc[매집순매수3등]['순매수']
            
            #-------------------------------#
            # 매물대 구간 누적 수급 구하기
            누적수급데이터 = all4[all4['일자'] == 매물대일자] # 매물대 일자의 수급 가져오기
            del 누적수급데이터['일자'] # 일자는 빼기
            누적수급데이터 = 누적수급데이터.transpose() # 행열 교환하기
            누적수급데이터.columns = ['매물대1누적'] # 컬럼명 지정하기
            누적수급데이터['누적랭크'] = 누적수급데이터['매물대1누적'].rank(method = 'first', ascending = False) # 랭크달기
            누적1등 = 누적수급데이터[누적수급데이터['누적랭크'] == 1.0].index[0]
            누적량1 = 누적수급데이터.loc[누적1등]['매물대1누적']
            누적2등 = 누적수급데이터[누적수급데이터['누적랭크'] == 2.0].index[0]
            누적량2 = 누적수급데이터.loc[누적2등]['매물대1누적']
            누적3등 = 누적수급데이터[누적수급데이터['누적랭크'] == 3.0].index[0]
            누적량3 = 누적수급데이터.loc[누적3등]['매물대1누적']
            #-------------------------------#
            # append 하기
            data_add.append(매물대강도)
            data_add.append(매물대대비)
            data_add.append(매수1등)
            data_add.append(매수량1)
            data_add.append(매수2등)
            data_add.append(매수량2)
            data_add.append(순매수1등)
            data_add.append(순매수량1)
            data_add.append(순매수2등)
            data_add.append(순매수량2)
            data_add.append(순매수3등)
            data_add.append(순매수량3)
            data_add.append(매집매수1등)
            data_add.append(매집매수량1)
            data_add.append(매집매수2등)
            data_add.append(매집매수량2)
            data_add.append(매집순매수1등)
            data_add.append(매집순매수량1)
            data_add.append(매집순매수2등)
            data_add.append(매집순매수량2)
            data_add.append(매집순매수3등)
            data_add.append(매집순매수량3)
            data_add.append(누적1등)
            data_add.append(누적량1)
            data_add.append(누적2등)
            data_add.append(누적량2)
            data_add.append(누적3등)
            data_add.append(누적량3)
            '''

#conn = test2_cls()
#conn.exe_test2("삼성카드", "20160128", 1, 2, 20, 2, 60, 120)